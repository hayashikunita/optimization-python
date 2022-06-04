# 参照
# https://gammasoft.jp/support/how-to-use-openpyxl-for-excel-file/

# openpyxlをインポート
import openpyxl

# Excelファイルの読み込み
wb = openpyxl.load_workbook("Sample.xlsx")


# Excelファイルの新規作成
wb = openpyxl.Workbook()


# 上書き保存（読み込んだのと同じ名前を指定）
# wb.save("Sample.xlsx")

# 別名保存（読み込んだのと別名を指定）
wb.save("Sample2Sample2.xlsx")


# 名前が「Sheet1」のシート
# ws = wb["Sheet1"]

# 先頭のシート（注意：インデックス番号は0から始まる）
# ws = wb.worksheets[0]

# # インデックス番号の確認
# >>> wb.index(ws)

# シート名のリスト
# >>> wb.sheetnames
# ['Sheet1', 'Sheet2', 'Sheet3']


# シート名の確認
# >>> ws.title
# 'Sheet1'

# シート名の変更
# >>> ws.title = "SheetOne"
# >>> wb.sheetnames
# ['SheetOne', 'Sheet2', 'Sheet3']

# >>> ws.title ="Sheet1"
# >>> wb.sheetnames
# ['Sheet1', 'Sheet2', 'Sheet3']

# 「Sheet4」を末尾に追加
# >>> ws4 = wb.create_sheet(title="Sheet4")
# >>> wb.sheetnames
# ['Sheet1', 'Sheet2', 'Sheet3', 'Sheet4']


# 「New Sheet」を左から３つ目の位置に追加
# >>> ws_new = wb.create_sheet(title="New Sheet", index=2)
# >>> wb.sheetnames
# ['Sheet1', 'Sheet2', 'New Sheet', 'Sheet3', 'Sheet4']


# 「Sheet2」のコピーをシートの最後に追加
# >>> ws2_copy = wb.copy_worksheet(wb["Sheet2"])
# >>> wb.sheetnames
# ['Sheet1', 'Sheet2', 'New Sheet', 'Sheet3', 'Sheet4', 'Sheet2 Copy']
# コピーしたシートは末尾に「 Copy」が付く

# 「Sheet2」のコピーを削除
# >>> wb.remove(ws2_copy)
# >>> wb.sheetnames
# ['Sheet1', 'Sheet2', 'New Sheet', 'Sheet3', 'Sheet4']

# 末尾のシートを削除
# >>> wb.remove(wb.worksheets[-1])
# >>> wb.sheetnames
# ['Sheet1', 'Sheet2', 'New Sheet', 'Sheet3']



# アドレス「A1」のセル
# >>> c1 = ws["A1"]
# >>> c1
# <Cell 'Sheet1'.A1>

# 「A1」を行列の番号で取得（注意：１から始まる）
# >>> c1 = ws.cell(row=1, column=1)

# キーワード（row=, column=）は省略も可能
# >>> c1 = ws.cell(1, 1)

# 「A1:C3」の範囲
# >>> rng1 = ws["A1:C3"]

# セルの文字列間をコロンで繋いでも可
# >>> rng1 = ws["A1":"C3"]

# １行分のセルのタプルを要素とするタプルになっている
# >>> rng1
# ((<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>), 
# (<Cell 'Sheet1'.A2>, <Cell 'Sheet1'.B2>, <Cell 'Sheet1'.C2>),
#  (<Cell 'Sheet1'.A3>, <Cell 'Sheet1'.B3>, <Cell 'Sheet1'.C3>))

# 範囲の１行分のセルはインデックスで取得できる
# >>> rng1[0]
# (<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>)


# シートの１行目（注意：１から始まる）
# >>> row1 = ws[1]

# １行分のセルがタプルになっている（この例はD列までデータがある場合）
# >>> row1
# (<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>, <Cell 'Sheet1'.D1>)

# １つのセルはインデックスで取得できる
# >>> row1[0]
# <Cell 'Sheet1'.A1>

# アドレス
# >>> c1.coordinate
# 'A1'

# 行番号
# >>> c1.row
# 1

# 列番号
# >>> c1.column
# 1

# 列アルファベット
# >>> c1.column_letter
# 'A'

# セルの値の読み取り
# >>> val1 = c1.value
# >>> val1
# 10000

# セルの値の書き込み
# >>> c1.value = 12000
# >>> c1.value
# 12000

# セルの数式の書き込み
# >>> c_sum = ws["C6"]
# >>> c_sum.value = "=SUM(C1:C5)"

# 数値（少数点以下の桁数指定）
# >>> ws["B2"].number_format = "0.00"

# 日付データの書式設定
# >>> ws["A2"].number_format = "yyyy年mm月dd日"

# フォントの設定（まずFontをインポートする）
# >>> from openpyxl.styles import Font

# 太字のON/OFF（bold=）、斜字のON/OFF（italic=）
# >>> ws["C2"].font = Font(bold=True, italic=True)

